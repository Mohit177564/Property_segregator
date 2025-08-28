#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from io import BytesIO
import tempfile

import streamlit as st
from openpyxl import load_workbook

from segregate_by_customer_code import segregate

st.set_page_config(
    page_title="Excel Customer Segregator",
    page_icon="🗂️",
    layout="wide",
)

st.title("🗂️ Excel Segregator by Customer Code")
st.caption("Upload an Excel workbook, select options, and download a new workbook with one sheet per customer code.")

with st.sidebar:
    st.header("How it works")
    st.write(
        """
        1. Upload a .xlsx file with your data.
        2. Choose the sheet, the column containing the customer codes, and how many header rows to preserve.
        3. Click Segregate to generate a workbook with a sheet per code.
        """
    )

uploaded = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx", "xlsm"], accept_multiple_files=False)

sheet_choice: str | None = None
column_spec: str = "F"
header_rows: int = 8

if uploaded is not None:
    data = uploaded.read()
    bio = BytesIO(data)
    try:
        wb = load_workbook(bio, read_only=True, data_only=True)
        sheetnames = wb.sheetnames
    except Exception as e:
        st.error(f"Could not read workbook: {e}")
        st.stop()

    c1, c2, c3 = st.columns(3)
    with c1:
        sheet_choice = st.selectbox("Worksheet", options=sheetnames, index=0)
    with c2:
        column_spec = st.text_input("Customer code column", value="F", help="Letter (e.g., F), 1-based index (e.g., 6), or header label.")
    with c3:
        header_rows = st.number_input("Header rows to preserve", min_value=0, max_value=100, value=8, step=1)

    st.divider()
    if st.button("Segregate", type="primary"):
        with st.spinner("Processing…"):
            try:
                # Persist upload to a temp file so openpyxl can work with a filesystem path
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
                    tmp_in.write(data)
                    tmp_in_path = Path(tmp_in.name)

                # Create a temp output path for the result
                tmp_out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp_out_path = Path(tmp_out.name)
                tmp_out.close()

                result = segregate(
                    input_path=tmp_in_path,
                    output_path=tmp_out_path,
                    sheet=sheet_choice,
                    column_spec=column_spec,
                    header_rows=int(header_rows),
                )

                out_bytes = Path(result).read_bytes()
                st.success("Segregation complete!")
                st.download_button(
                    label="Download segregated workbook",
                    data=out_bytes,
                    file_name=f"{Path(uploaded.name).stem}_segregated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Failed to segregate: {e}")
else:
    st.info("Upload an Excel file to begin.")
