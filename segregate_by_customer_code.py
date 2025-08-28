#!/usr/bin/env python3
"""
Segregate rows in an Excel file into separate sheets by customer code.

Default behavior:
- Reads the first sheet of the input workbook (default: ./test.xlsx)
- Uses column F as the customer code column
- Creates a new workbook named <input>_segregated.xlsx with one sheet per code

Usage examples:
  python3 segregate_by_customer_code.py --input test.xlsx --column F
  python3 segregate_by_customer_code.py --input test.xlsx --column "Customer Code"
  python3 segregate_by_customer_code.py --input test.xlsx --sheet "Sheet1" --output output.xlsx

Notes:
- Requires pandas and openpyxl (see requirements.txt)
- Sheet names are sanitized to be Excel‑safe and truncated to 31 chars if needed
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any, Iterable, Set

import pandas as pd
from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook, Workbook
from copy import copy as _copy


def col_letter_to_index(letter: str) -> int:
    """Convert Excel column letter (e.g., 'F') to 0-based index."""
    return column_index_from_string(letter.upper()) - 1


def resolve_key_column(column_spec: str, columns: Iterable[Any]) -> Any:
    """Resolve the column in the DataFrame to group by, based on a user spec.

    column_spec can be:
      - Excel letter like 'F'
      - 1-based index like '6'
      - Exact or case-insensitive column header name

    Returns the DataFrame column name to use for grouping.
    """
    cols = list(columns)
    s = str(column_spec).strip()

    # Numeric (1-based)
    if s.isdigit():
        idx = int(s) - 1
        if idx < 0 or idx >= len(cols):
            raise ValueError(f"Column index {s} is out of range (1..{len(cols)})")
        return cols[idx]

    # Letter(s)
    if re.fullmatch(r"[A-Za-z]+", s):
        idx = col_letter_to_index(s)
        if idx < 0 or idx >= len(cols):
            raise ValueError(f"Column letter {s} resolves outside available columns (len={len(cols)})")
        return cols[idx]

    # Header name (exact or case-insensitive)
    if s in cols:
        return s
    for c in cols:
        if str(c).strip().lower() == s.lower():
            return c

    raise ValueError(
        f"Could not resolve column '{column_spec}'. Available columns: {cols}"
    )


def excel_safe_sheet_name(name: Any, used: Set[str]) -> str:
    """Make a value safe to use as an Excel sheet name and ensure uniqueness.

    - Replaces invalid characters : \\ / ? * [ ] with '_'
    - Truncates to 31 characters (Excel limit)
    - Ensures uniqueness by appending _1, _2, ... if needed
    """
    base = str(name) if pd.notna(name) else "Blank"
    # Replace invalid chars
    base = re.sub(r"[:\\/\?\*\[\]]", "_", base)
    if not base:
        base = "Sheet"
    # Truncate to 31
    base = base[:31]

    candidate = base
    i = 1
    while candidate in used or candidate == "":
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    used.add(candidate)
    return candidate


def copy_cell_style(src, dst) -> None:
    """Copy common cell style attributes from src to dst."""
    try:
        if getattr(src, "has_style", False):
            if src.font is not None:
                dst.font = _copy(src.font)
            if src.fill is not None:
                dst.fill = _copy(src.fill)
            if src.border is not None:
                dst.border = _copy(src.border)
            if src.alignment is not None:
                dst.alignment = _copy(src.alignment)
            if src.protection is not None:
                dst.protection = _copy(src.protection)
            # number_format is a string
            dst.number_format = src.number_format
    except TypeError:
        # Fallback: ignore style copy if objects are not hashable in this environment
        dst.number_format = src.number_format


def copy_header_and_layout(ws_src, ws_dst, header_rows: int) -> None:
    """Copy the top header_rows (values + styles), merges, widths, row heights and freeze panes."""
    max_col = ws_src.max_column

    # Copy column widths
    for key, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[key].width = dim.width

    # Copy header cells and row heights
    for r in range(1, header_rows + 1):
        if ws_src.row_dimensions[r].height is not None:
            ws_dst.row_dimensions[r].height = ws_src.row_dimensions[r].height
        for c in range(1, max_col + 1):
            s = ws_src.cell(row=r, column=c)
            d = ws_dst.cell(row=r, column=c, value=s.value)
            copy_cell_style(s, d)

    # Copy merged cells that intersect headers
    for rng in ws_src.merged_cells.ranges:
        if rng.min_row <= header_rows:
            ws_dst.merge_cells(str(rng))

    # Freeze panes below the header
    ws_dst.freeze_panes = f"A{header_rows + 1}"


def resolve_key_column_index(ws, column_spec: str, header_rows: int) -> int:
    """Resolve to a 1-based column index on the openpyxl worksheet.

    Accepts letter (e.g., 'F'), 1-based index ('6'), or a header label searched
    within the first `header_rows` rows (case-insensitive).
    """
    s = str(column_spec).strip()

    if s.isdigit():
        return int(s)

    if re.fullmatch(r"[A-Za-z]+", s):
        return column_index_from_string(s)

    # Search header area for a matching label
    s_lower = s.lower()
    for r in range(1, header_rows + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip().lower() == s_lower:
                return c

    raise ValueError(f"Could not resolve column '{column_spec}' in header area")


def segregate(
    input_path: Path,
    output_path: Path | None = None,
    sheet: str | int | None = None,
    column_spec: str = "F",
    header_rows: int = 8,
) -> Path:
    # Load source workbook and worksheet
    wb_src = load_workbook(input_path)
    if sheet is None:
        ws_src = wb_src.worksheets[0]
    elif isinstance(sheet, int):
        ws_src = wb_src.worksheets[sheet]
    else:
        ws_src = wb_src[sheet]

    max_row = ws_src.max_row
    max_col = ws_src.max_column
    if max_row <= header_rows:
        raise ValueError("The input sheet has no data rows below the header.")

    key_col = resolve_key_column_index(ws_src, column_spec, header_rows)

    # Prepare output path
    if output_path is None:
        output_path = input_path.with_name(f"{input_path.stem}_segregated{input_path.suffix}")

    # Collect row indices per code
    groups: dict[str, list[int]] = {}
    data_start = header_rows + 1
    for r in range(data_start, max_row + 1):
        code_val = ws_src.cell(row=r, column=key_col).value
        if code_val is None or str(code_val).strip() == "":
            continue
        code_key = str(code_val)
        groups.setdefault(code_key, []).append(r)

    if not groups:
        raise ValueError("No customer codes found in the specified column.")

    # Write output workbook
    wb_out = Workbook()
    # Remove default sheet
    default_ws = wb_out.active
    wb_out.remove(default_ws)

    used_names: Set[str] = set()

    for code, row_indices in groups.items():
        ws_dst = wb_out.create_sheet(title=excel_safe_sheet_name(code, used_names))
        # Copy header + layout
        copy_header_and_layout(ws_src, ws_dst, header_rows)

        out_r = data_start
        for r in row_indices:
            # Copy row height if present
            if ws_src.row_dimensions[r].height is not None:
                ws_dst.row_dimensions[out_r].height = ws_src.row_dimensions[r].height
            for c in range(1, max_col + 1):
                s = ws_src.cell(row=r, column=c)
                d = ws_dst.cell(row=out_r, column=c, value=s.value)
                copy_cell_style(s, d)
            out_r += 1

    wb_out.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Segregate Excel rows into separate sheets by customer code.")
    parser.add_argument("--input", "-i", type=Path, default=Path("test.xlsx"), help="Path to the source .xlsx file (default: ./test.xlsx)")
    parser.add_argument("--sheet", "-s", help="Sheet name or 0-based index to read (default: first sheet)")
    parser.add_argument("--column", "-c", default="F", help="Customer code column (letter like F, 1-based index, or header name). Default: F")
    parser.add_argument("--output", "-o", type=Path, help="Output .xlsx path (default: <input>_segregated.xlsx)")
    parser.add_argument("--header-rows", type=int, default=8, help="Number of header rows at the top to copy verbatim. Default: 8")

    args = parser.parse_args()

    sheet_arg: str | int | None
    if args.sheet is None:
        sheet_arg = None
    else:
        # Try to parse as int index; else treat as name
        try:
            sheet_arg = int(args.sheet)
        except ValueError:
            sheet_arg = args.sheet

    out = segregate(
        input_path=args.input,
        output_path=args.output,
        sheet=sheet_arg,
        column_spec=args.column,
        header_rows=args.header_rows,
    )
    print(f"Created: {out}")


if __name__ == "__main__":
    main()
